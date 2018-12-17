from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime
from datetime import date

class ExcelManager(object):


    def __init__(self, targhet):
        self.dataA5 = date(2018, 12, 16)
        self.targhet = targhet
        self.file_name = targhet + '.xlsx'
        self.wb = self.load_excel(self.file_name)

    def load_excel(self, player):
        try:
            wb = load_workbook(player)
        except Exception as e:
            template = load_workbook('template.xlsx')
            template.save(player)
            template.close()
            wb = load_workbook(player)
        return wb

    def select_sheet(self, wb, coords, isMoon):
        sheet_name = coords.replace(":", "-");
        if isMoon:
            sheet_name = sheet_name + ' (Luna)'

        try:
            ws = wb.get_sheet_by_name(sheet_name)
        except Exception as e:
            ws = wb.get_sheet_by_name("Empty")
            ws = wb.copy_worksheet(ws)
            ws.title = sheet_name
        return ws

    def findRow(self):
        now = datetime.datetime.now()
        oggi = date(now.year, now.month, now.day)
        delta = oggi - self.dataA5
        return delta.days + 5


    def findCol(self, ws):
        now = datetime.datetime.now()
        time = str(now.hour).zfill(2) + str(now.minute).zfill(2) + '00'
        nOra = int(time)

        i = 2

        while i < 50:
            i = i + 1
            _min_t = int(str(ws[get_column_letter(i) + str(4)] .value).replace(':', ''))
            _max_t = int(str(ws[get_column_letter(i+1) + str(4)] .value).replace(':', ''))

            if (nOra > _min_t) and (nOra < _max_t):
                return get_column_letter(i)
        return get_column_letter(i)

    def scrivi(self, ws, row, col, time):
        val = ws[col + str(row)].value

        if val == None:
            ws[col + str(row)] = time
        else:
            if time < int(val):
                ws[col + str(row)] = time

    def write_time(self, coords, isMoon, time):
        if(time == ">60"):
            time = 999

        ws = self.select_sheet(self.wb, coords, isMoon)
        row = self.findRow()
        col = self.findCol(ws)
        self.scrivi(ws, row, col, time);

        ws = self.wb.get_sheet_by_name("Attivita Generale")
        self.scrivi(ws, row, col, time);

        self.wb.save(self.file_name)

if __name__ == '__main__':
    excelManager = ExcelManager("Prova")
    excelManager.write_time("Ciao", True, 10)
